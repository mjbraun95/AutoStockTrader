from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Color
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import datetime
import ISMScraper
import requests

#DEFINE CELL COLORS
DARKRED = 'FFC00000'
RED = 'FFFF0000'
ORANGE = 'FFFFC000'
YELLOW = 'FFFFFF00'
GREEN = 'FF00B050'
BLUE = 'FF0070C0'
LIGHTBLUE = 'FF00B0F0'
PURPLE = 'FF7030A0'
WHITE = '00000000'
OTHERWHITE = 'FFFFFFFF'
black = (0,0,0)
white = (255,255,255)
red = (255,0,0)
green = (0,255,0)
blue = (0,0,255)
grey = (128,128,128)


debug = True

Sectors_xlsx = "Sectors.xlsx"
Sectors_Directory = "C:/Users/mattb/Dropbox/Github Repositories/AutoStockTrader/Spreadsheets/PMI_NMI/{}".format( Sectors_xlsx ) 

PMI_IndustriesArray = ["(Machinery)","(Computer & Electronic Products)","(Paper Products)",
                    "(Apparel, Leather & Allied Products)","(Printing & Related Support Activities)",
                    "(Primary Metals)","(Nonmetallic Mineral Products)","(Petroleum & Coal Products)",
                    "(Plastics & Rubber Products)","(Miscellaneous Manufacturing)",
                    "(Food, Beverage & Tobacco Products)","(Furniture & Related Products)",
                    "(Transportation Equipment)","(Chemical Products)","(Fabricated Metal Products)",
                    "(Electrical Equipment, Appliances & Components)","(Textile Mills)","(Wood Products)"]
NMI_IndustriesArray = ["(Retail Trade)","(Utilities)","(Arts, Entertainment Recreation)",
                    "(Other Services)","(Healthcare and Social Assistance)","(Food and Accomodations)",
                    "(Finance and Insurance)","(Real Estate, Renting and Leasing)",
                    "(Transport and Warehouse)","(Mining)","(Wholesale)","(Public Admin)",
                    "(Professional, Science and Technology Services)","(Information)","(Education)",
                    "(Management)","(Construction)","(Agriculture, Forest, Fishing and Hunting)"]
PMI_SectorsArray = ["NEW ORDERS","PRODUCTION","EMPLOYMENT","SUPPLIER DELIVERIES","INVENTORIES",
                    "CUSTOMER INVENTORIES","PRICES","BACKLOG OF ORDERS","EXPORTS","IMPORTS"]
NMI_SectorsArray = ["ISM NON-MANUFACTURING","BUSINESS ACTIVITY","NEW ORDERS","EMPLOYMENT","SUPPLIER DELIVERIES","INVENTORIES",
                    "PRICES","BACKLOG OF ORDERS","EXPORTS","IMPORTS"]

IndustryCommentsInitColumn = 68
IndustryCommentsInitRow = 5

IndustryTrendInitColumn = 71
IndustryTrendInitRow = 6

class cellEditor:
    def __init__(self,xlsx,sheetIndex,row=1,column=65):
        self.xlsx = xlsx                                    #e.g. "MultipleTimingSP500PMI_xlsx.xlsx"
        self.workbook = load_workbook(filename = self.xlsx)                    #e.g. load_workbook(filename = "MultipleTimingSP500PMI_xlsx.xlsx")
        if debug == True:
            print("self.workbook: {}".format(self.workbook))
        self.sheet = self.workbook.worksheets[sheetIndex]                                  #e.g. MyDistribution
        if debug == True:
            print("self.sheet: {}".format(self.sheet))
        self.row = row                                      #e.g. 7
        self.column = column                                #e.g. A
        self.coordinates = "{}{}".format(chr(self.column),self.row)   #e.g. 'A7'
        self.cell = self.sheet["{}".format(self.coordinates)]  #e.g. MyDistribution['A7']
        self.value = self.cell.value                        #e.g. "this is the text contained in the current cell"

    #Style change functions
    def changeCellTextColor(self,color):
        self.cell.font = Font(color=color)
        self.saveWorkbook()
        if (debug == True):
            self.printCellFont()
    def changeCellBackgroundColor(self,color):
        self.cell.fill = color
        if (debug == True):
            self.printCellFont()
    def changeCellFont(self, name = "Calibri (Body)", sz = 11, bold = False, italic = False, underline = False, color = black):
        self.cell.font = Font(name=name, sz=sz, bold=bold, italic=italic, underline=underline, color=color)
    #Printing functions
    def printCellCoordinates(self):
        print("Current cell coordinates: {}".format(self.coordinates))
    def printCellValue(self):
        print("Cell {} value is now: {}".format(self.coordinates,self.cell.value))
    def printCellBackgroundColor(self):
        print("Cell {} background color is now: {}".format(self.coordinates,self.cell.fill))
    def printCellFont(self):
        print("Cell {} font is now: {}".format(self.coordinates,self.cell.font))

    #Save xlsx file
    def saveWorkbook(self):
        self.workbook.save(filename = "{}".format(self.xlsx))
        print("{} Saved!".format( self.xlsx ))

    #Coordinate change functions
    def cellCoordAdjust(self):
        self.coordinates = "{}{}".format(chr(self.column),self.row)
        self.cell = self.sheet["{}".format(self.coordinates)]
        self.value = self.cell.value
        self.saveWorkbook()
    def changeRow(self,value):
        self.row = value
        self.cellCoordAdjust()
        if (debug == True):
            self.printCellCoordinates()
    def changeColumn(self,value):
        self.column = value
        self.cellCoordAdjust()
        if (debug == True):
            self.printCellCoordinates()

    #Value change functions
    def changeCellValue(self,text):
        self.cell.value = text
        # self.saveWorkbook()
        if (debug == True):
            self.printCellValue()
        

    def updateComments( self, IndustriesArray, IndustryCommentsCurrentColumn = IndustryCommentsInitColumn, IndustryCommentsCurrentRow = IndustryCommentsInitRow ):
        print("Updating {} comments...".format( self.sheet ))
        self.column = IndustryCommentsCurrentColumn
        self.row = IndustryCommentsCurrentRow
        ISM_CommentList = ISMScraper.grabISMcomments(ISM_Page)
        for noteIndex, note in enumerate(ISM_CommentList):
            for industryIndex, industry in enumerate(IndustriesArray):
                if debug == True:
                    print("Searching for {} in note {}!".format( industry.strip("()"), noteIndex))
                if note.text.find(industry) != -1:
                    if debug == True:
                        print("note.text.find(industry): {}".format(note.text.find(industry)))
                        print("Found {} in note {}!".format( industry.strip("()"), noteIndex))

                    IndustryCommentsCurrentRow += (2*industryIndex)
                    self.row = IndustryCommentsCurrentRow
                    self.cellCoordAdjust()
                    self.changeCellValue(ISM_CommentList[noteIndex].text)

                    IndustryCommentsCurrentRow = IndustryCommentsInitRow
                    self.row = IndustryCommentsInitRow
                    self.cellCoordAdjust()
                    break

            print("ERROR: NO MATCHING INDUSTRY")
        print("Program executed successfully! :) Saving...")
        self.saveWorkbook()
        return 0

    def updateRankings( self ):
        IndustryPMIRankingsInitColumn = 71
        IndustryNMIRankingsInitColumn = 70
        IndustryRankingsInitRow = 5
        print("Updating PMI rankings...")
        ISM_Type = ISMScraper.grabType(ISM_Page)
        if ISM_Type == "PMI":
            thisInitColumn = IndustryPMIRankingsInitColumn
            ISM_RankingListArray = [None] * 10
            thisSectorsArray = PMI_SectorsArray
            thisIndustryArray = PMI_IndustriesArray
        elif ISM_Type == "NMI":
            thisInitColumn = IndustryNMIRankingsInitColumn
            ISM_RankingListArray = [None] * 11
            thisSectorsArray = NMI_SectorsArray
            thisIndustryArray = NMI_IndustriesArray
        rankDictionaryArray = ISMScraper.grabISMrankings( ISM_Page )

        self.column = thisInitColumn
        for rankDictIndex, rankDictionary in enumerate(rankDictionaryArray):
            for currentIndustry, rank in rankDictionary.items():
                self.row = IndustryRankingsInitRow
                # Searching for correct row to put rank into
                for industryIndex, industry in enumerate( thisIndustryArray ):
                    if industry.find(currentIndustry) != -1:
                        self.row += 2*industryIndex
                        print("Found {}!".format(currentIndustry))
                        self.cellCoordAdjust()
                        self.changeCellValue( rank )
                        break
                    
                
            self.column += 1



        ArrayIndex = 0
        print("len(ISM_RankingListArray): {}".format( len (ISM_RankingListArray) ) )
        
        print("Program executed successfully! :) Saving...")
        self.saveWorkbook()
        return 0

def updateSheetNames( xlsx ):
    print("\nGiving sheet variables to each sheet...")
    if debug == True:
        print("Sheet names for {}:".format( xlsx ))
        for sheet_i, sheet in enumerate(xlsx.worksheets):
            print("Sheet {}: {}".format( sheet_i, sheet ))
    print("Done giving sheet variables to the sheet!")
    return 0


PMI_Editor = cellEditor( Sectors_xlsx, 1 )
NMI_Editor = cellEditor( Sectors_xlsx, 2 )

# URL = input("Enter URL Here: ")
ISM_Page = requests.get("https://www.instituteforsupplymanagement.org/ISMReport/MfgROB.cfm?SSO=1") #PMI
# ISM_Page = requests.get("https://www.instituteforsupplymanagement.org/ISMReport/NonMfgROB.cfm?SSO=1") #NMI
if ISM_Page.status_code != 200:
    print("ERROR: status code: {}".format(ISM_Page.status_code))
    exit()
else:
    print("Request successful!")

command = ""
PMI_Editor.updateRankings( )
# while command != 'q':
#     command = input("What would you like to do? ('q'=quit, 'uc'=Update Comments, 'ur'=Update Rankings, ")
#     if command == 'uc':
#         PMI_Editor.updateComments( PMI_IndustriesArray )
#     elif command == 'ur':
#         PMI_Editor.updateRankings( PMI_IndustriesArray )
print("Exiting")