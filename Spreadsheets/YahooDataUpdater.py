from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Color
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import datetime
import YahooScraper

debug = True

#DEFINE CELL COLORS
DARKRED = 'FFC00000'
RED = 'FFFF0000'
ORANGE = 'FFFFC000'
YELLOW = 'FFFFFF00'
GREEN = 'FF00B050'
BLUE = 'FF0070C0'
LIGHTBLUE = 'FF00B0F0'
PURPLE = 'FF7030A0'
WHITE = '00000000'
OTHERWHITE = 'FFFFFFFF'

black = (0,0,0)
white = (255,255,255)
red = (255,0,0)
green = (0,255,0)
blue = (0,0,255)
grey = (128,128,128)

excelFile = input("Enter save name for spreadsheet (default is 'Test1.xlsx'): ")
if excelFile == "" or excelFile == " ":
    excelFile = "Test1.xlsx"
# excelDirectory = "C:\Users\mattb\Dropbox\Github Repositories\AutoStockTrader\Spreadsheets\Test1.xlsx" 



class cellEditor:
    def __init__(self,row,column,sheet,distribution):
        self.excelFile = excelFile                          #e.g. "MultipleTimingSP500Distribution1.xlsx"
        self.distribution = distribution                    #e.g. load_workbook(filename = "MultipleTimingSP500Distribution1.xlsx")
        if debug == True:
            print("self.distribution: {}".format(self.distribution))
        self.sheet = sheet                                  #e.g. MyDistribution
        self.row = row                                      #e.g. 7
        self.column = column                                #e.g. A
        self.coordinates = "{}{}".format(chr(self.column),self.row)   #e.g. 'A7'
        if debug == True:
            print("self.coordinates: {}".format(self.coordinates))
        self.cell = self.sheet["{}".format(self.coordinates)]  #e.g. MyDistribution['A7']
        self.value = self.cell.value                        #e.g. "this is the text contained in the current cell"

    # def cellEditor()

    #Printing functions
    def printCellCoordinates(self):
        print("Current cell coordinates: {}".format(self.coordinates))
    def printCellValue(self):
        print("Cell {} value is now: {}".format(self.coordinates,self.cell.value))
    def printCellBackgroundColor(self):
        print("Cell {} background color is now: {}".format(self.coordinates,self.cell.fill))
    def printCellFont(self):
        print("Cell {} font is now: {}".format(self.coordinates,self.cell.font))

    #Save xlsx file
    def saveToDistribution(self):
        self.distribution.save(filename = "{}".format(self.excelFile))

    #Coordinate change functions
    def cellCoordAdjust(self):
        self.coordinates = "{}{}".format(chr(self.column),self.row)
        self.cell = self.sheet["{}".format(self.coordinates)]
        self.value = self.cell.value
        self.saveToDistribution()
    def changeRow(self,value):
        self.row = value
        self.cellCoordAdjust()
        if (debug == True):
            self.printCellCoordinates()
    def changeColumn(self,value):
        self.column = value
        self.cellCoordAdjust()
        if (debug == True):
            self.printCellCoordinates()

    #Value change functions
    def changeCellValue(self,text):
        self.cell.value = text
        # self.saveToDistribution()
        if (debug == True):
            self.printCellValue()

    #Style change functions
    def changeCellTextColor(self,color):
        self.cell.font = Font(color=color)
        self.saveToDistribution()
        if (debug == True):
            self.printCellFont()
    def changeCellBackgroundColor(self,color):
        self.cell.fill = color
        if (debug == True):
            self.printCellFont()
    def changeCellFont(self, name = "Calibri (Body)", sz = 11, bold = False, italic = False, underline = False, color = black):
        self.cell.font = Font(name=name, sz=sz, bold=bold, italic=italic, underline=underline, color=color)

currentDate = datetime.date(2019, 1, 6).isoformat()

xTextList = 0
currentText = 0

Distribution1 = load_workbook(filename = excelFile)

# Giving variable names to each sheet...
print("\nGiving title variables to each sheet...")
# print("Distribution1.sheetnames: {}".format(Distribution1.sheetnames))
# print("Distribution1.sheetnames[1]: {}".format(Distribution1.sheetnames[1]))

Sheet1 = Distribution1.worksheets[0]
# MyDistributionTitle = Distribution1.worksheets[0]
# DistributionTemplateTitle = Distribution1.worksheets[1]
# VolCalculationsTitle = Distribution1.worksheets[2]
# ChartPornTitle = Distribution1.worksheets[3]
# SectorScreenerTitle = Distribution1.worksheets[4]
# VIXImpliedSP500MovesTitle = Distribution1.worksheets[5]
# GDP_SP500_RelationshipTitle = Distribution1.worksheets[6]
# ValueChainTitle = Distribution1.worksheets[7]
# WatchListTitle = Distribution1.worksheets[8]

print("Sheet1: {}".format(Sheet1))

# MyDistributionTitle = "MyDistributionTitle"
# DistributionTemplateTitle = "DistributionTemplateTitle"
# VolCalculationsTitle = "VolCalculationsTitle"
# ChartPornTitle = "ChartPornTitle"
# SectorScreenerTitle = "SectorScreenerTitle"
# VIXImpliedSP500MovesTitle = "VIXImpliedSP500MovesTitle"
# GDP_SP500_RelationshipTitle = "GDP_SP500_RelationshipTitle"
# ValueChainTitle = "ValueChainTitle"
# WatchListTitle = "WatchListTitle"

print("Done giving variable names to each sheet!")

# print("\nGiving sheet variables to each sheet...")
# MyDistributionSheet = Distribution1[MyDistributionTitle]
# DistributionTemplateSheet = Distribution1[DistributionTemplateTitle]
# VolCalculationsSheet = Distribution1[VolCalculationsTitle]
# ChartPornSheet = Distribution1[ChartPornTitle]
# SectorScreenerSheet = Distribution1[SectorScreenerTitle]
# VIXImpliedSP500MovesSheet = Distribution1[VIXImpliedSP500MovesTitle]
# GDP_SP500_RelationshipSheet = Distribution1[GDP_SP500_RelationshipTitle]
# ValueChainSheet = Distribution1[ValueChainTitle]
# WatchListSheet = Distribution1[WatchListTitle]
# print("Done giving sheet variables to each sheet!")

print("\nUpdating DistributionTemplateSheet...")
# Input o/h/l/c/a/v coordinate values
initRow = 7
dateColumn = 65
openColumn = 66
highColumn = 67
lowColumn = 68
closeColumn = 69
adjCloseColumn = 70
volumeColumn = 71

dateEditor = cellEditor(initRow,dateColumn,Sheet1,Distribution1)
openEditor = cellEditor(initRow,openColumn,Sheet1,Distribution1)
highEditor = cellEditor(initRow,highColumn,Sheet1,Distribution1)
lowEditor = cellEditor(initRow,lowColumn,Sheet1,Distribution1)
closeEditor = cellEditor(initRow,closeColumn,Sheet1,Distribution1)
adjCloseEditor = cellEditor(initRow,adjCloseColumn,Sheet1,Distribution1)
volumeEditor = cellEditor(initRow,volumeColumn,Sheet1,Distribution1)

numRows = int(input("Write to how many rows?"))
currentRow = initRow

#Grabs Finance Table Data from the YahooScraper.py file
yahooFinanceDataRows = YahooScraper.grabYahooFinanceDataRows()


# currentRow = 0
while currentRow < numRows:
    dateEditor.changeCellValue(yahooFinanceDataRows[currentRow-initRow][0])
    openEditor.changeCellValue(yahooFinanceDataRows[currentRow-initRow][1])
    highEditor.changeCellValue(yahooFinanceDataRows[currentRow-initRow][2])
    lowEditor.changeCellValue(yahooFinanceDataRows[currentRow-initRow][3])
    closeEditor.changeCellValue(yahooFinanceDataRows[currentRow-initRow][4])
    adjCloseEditor.changeCellValue(yahooFinanceDataRows[currentRow-initRow][5])
    volumeEditor.changeCellValue(yahooFinanceDataRows[currentRow-initRow][6])
    print("yahooFinanceDataRows[0][0]: {}".format(yahooFinanceDataRows[0][0]))
    # del yahooFinanceDataRows[currentRow-initRow][:]

    currentRow += 1

    # Update the cells' rows
    dateEditor.changeRow(currentRow)
    openEditor.changeRow(currentRow)
    highEditor.changeRow(currentRow)
    lowEditor.changeRow(currentRow)
    closeEditor.changeRow(currentRow)
    adjCloseEditor.changeRow(currentRow)
    volumeEditor.changeRow(currentRow)
    print("Done row {}!".format(currentRow))
    
print("Program executed successfully! :) Saving...")
openEditor.saveToDistribution()
print("Saved!")












# Distribution1.save(filename = "spreadsheet1.xlsx")
# print("closeColumn: {}".format(chr(closeColumn)))
# def __init__(self, name = volCalcRange[currentCloseCell].value, 
#     priority = (volCalcRange['{}{}'.format(chr(closeColumn-2),closeRow)].value), 
#     start = volCalcRange['{}{}'.format(chr(closeColumn-1),closeRow)].value, 
#     deadline = volCalcRange['{}{}'.format(chr(closeColumn+1),closeRow)].value,
#     repeating = volCalcRange[currentCloseCell].font.b, 
#     calendar = volCalcRange[currentCloseCell].font.i, 
#     color = volCalcRange[currentCloseCell].fill.start_color.index, 
#     topRow = closeRow, 
#     leftColumn = closeColumn, 
#     end = False):