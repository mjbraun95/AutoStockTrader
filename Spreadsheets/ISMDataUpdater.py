from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Color
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import datetime
import ISMScraper

debug = True

#DEFINE CELL COLORS
DARKRED = 'FFC00000'
RED = 'FFFF0000'
ORANGE = 'FFFFC000'
YELLOW = 'FFFFFF00'
GREEN = 'FF00B050'
BLUE = 'FF0070C0'
LIGHTBLUE = 'FF00B0F0'
PURPLE = 'FF7030A0'
WHITE = '00000000'
OTHERWHITE = 'FFFFFFFF'

black = (0,0,0)
white = (255,255,255)
red = (255,0,0)
green = (0,255,0)
blue = (0,0,255)
grey = (128,128,128)

# excelFile = input("Enter save name for spreadsheet (default is 'Test1.xlsx'): ")
# if excelFile == "" or excelFile == " ":
excelFile = "NewPMISectors2.xlsx"
# excelDirectory = "C:\Users\mattb\Dropbox\Github Repositories\AutoStockTrader\Spreadsheets\Test1.xlsx" 



class cellEditor:
    def __init__(self,row,column,sheet,distribution):
        self.excelFile = excelFile                          #e.g. "MultipleTimingSP500PMIExcelFile.xlsx"
        self.distribution = distribution                    #e.g. load_workbook(filename = "MultipleTimingSP500PMIExcelFile.xlsx")
        if debug == True:
            print("self.distribution: {}".format(self.distribution))
        self.sheet = sheet                                  #e.g. MyDistribution
        self.row = row                                      #e.g. 7
        self.column = column                                #e.g. A
        self.coordinates = "{}{}".format(chr(self.column),self.row)   #e.g. 'A7'
        if debug == True:
            print("self.coordinates: {}".format(self.coordinates))
        self.cell = self.sheet["{}".format(self.coordinates)]  #e.g. MyDistribution['A7']
        self.value = self.cell.value                        #e.g. "this is the text contained in the current cell"

    # def cellEditor()

    #Printing functions
    def printCellCoordinates(self):
        print("Current cell coordinates: {}".format(self.coordinates))
    def printCellValue(self):
        print("Cell {} value is now: {}".format(self.coordinates,self.cell.value))
    def printCellBackgroundColor(self):
        print("Cell {} background color is now: {}".format(self.coordinates,self.cell.fill))
    def printCellFont(self):
        print("Cell {} font is now: {}".format(self.coordinates,self.cell.font))

    #Save xlsx file
    def saveToDistribution(self):
        self.distribution.save(filename = "{}".format(self.excelFile))

    #Coordinate change functions
    def cellCoordAdjust(self):
        self.coordinates = "{}{}".format(chr(self.column),self.row)
        self.cell = self.sheet["{}".format(self.coordinates)]
        self.value = self.cell.value
        self.saveToDistribution()
    def changeRow(self,value):
        self.row = value
        self.cellCoordAdjust()
        if (debug == True):
            self.printCellCoordinates()
    def changeColumn(self,value):
        self.column = value
        self.cellCoordAdjust()
        if (debug == True):
            self.printCellCoordinates()

    #Value change functions
    def changeCellValue(self,text):
        self.cell.value = text
        # self.saveToDistribution()
        if (debug == True):
            self.printCellValue()

    #Style change functions
    def changeCellTextColor(self,color):
        self.cell.font = Font(color=color)
        self.saveToDistribution()
        if (debug == True):
            self.printCellFont()
    def changeCellBackgroundColor(self,color):
        self.cell.fill = color
        if (debug == True):
            self.printCellFont()
    def changeCellFont(self, name = "Calibri (Body)", sz = 11, bold = False, italic = False, underline = False, color = black):
        self.cell.font = Font(name=name, sz=sz, bold=bold, italic=italic, underline=underline, color=color)

currentDate = datetime.date(2019, 1, 6).isoformat()

xTextList = 0
currentText = 0

PMIExcelFile = load_workbook(filename = excelFile)

# Giving variable names to each sheet...
print("\nGiving title variables to each sheet...")
# print("PMIExcelFile.sheetnames: {}".format(PMIExcelFile.sheetnames))
# print("PMIExcelFile.sheetnames[1]: {}".format(PMIExcelFile.sheetnames[1]))

# PMIAnalysis = PMIExcelFile.worksheets[0]
Sectors = PMIExcelFile.worksheets[1]
# IndustryComments = PMIExcelFile.worksheets[2]

# print("IndustryComments: {}".format(IndustryComments))

print("\nGiving sheet variables to each sheet...")
SectorsSheet = PMIExcelFile["Sheet2"]
print("Done giving sheet variables to the sheet!")

# print("\nUpdating DistributionTemplateSheet...")
# Input o/h/l/c/a/v coordinate values
# initRow = 7
# dateColumn = 65
# openColumn = 66
# highColumn = 67
# lowColumn = 68
# closeColumn = 69
# adjCloseColumn = 70
# volumeColumn = 71
IndustriesArray = ["(Machinery)","(Computer & Electronic Products)","(Paper Products)",
                    "(Apparel, Leather & Allied Products)","(Printing & Related Support Activities)",
                    "(Primary Metals)","(Nonmetallic Mineral Products)","(Petroleum & Coal Products)",
                    "(Plastics & Rubber Products)","(Miscellaneous Manufacturing)",
                    "(Food, Beverage & Tobacco Products)","(Furniture & Related Products)",
                    "(Transportation Equipment)","(Chemical Products)","(Fabricated Metal Products)",
                    "(Electrical Equipment, Appliances & Components)","(Textile Mills)","(Wood Products)"]



# dateEditor = cellEditor(initRow,dateColumn,IndustryComments,PMIExcelFile)
# openEditor = cellEditor(initRow,openColumn,IndustryComments,PMIExcelFile)
# highEditor = cellEditor(initRow,highColumn,IndustryComments,PMIExcelFile)
# lowEditor = cellEditor(initRow,lowColumn,IndustryComments,PMIExcelFile)
# closeEditor = cellEditor(initRow,closeColumn,IndustryComments,PMIExcelFile)
# adjCloseEditor = cellEditor(initRow,adjCloseColumn,IndustryComments,PMIExcelFile)
# volumeEditor = cellEditor(initRow,volumeColumn,IndustryComments,PMIExcelFile)

# numRows = int(input("Write to how many rows?"))
# currentRow = initRow

#Grabs Finance Table Data from the ISMScraper.py file
ISMnotelist = ISMScraper.grabISMnotes()

IndustryCommentsInitColumn = 66
IndustryCommentsInitRow = 3
IndustryCommentsCurrentColumn = IndustryCommentsInitColumn
IndustryCommentsCurrentRow = IndustryCommentsInitRow

IndustryCommentsEditor = cellEditor(IndustryCommentsInitRow,IndustryCommentsInitColumn,Sectors,PMIExcelFile)

for noteIndex, note in enumerate(ISMnotelist):
    for industryIndex, industry in enumerate(IndustriesArray):
        if debug == True:
            print("Searching for {} in note {}!".format( industry.strip("()"), noteIndex))
        if note.text.find(industry) != -1:
            if debug == True:
                print("note.text.find(industry): {}".format(note.text.find(industry)))
                print("Found {} in note {}!".format( industry.strip("()"), noteIndex))

            IndustryCommentsCurrentRow += (2*industryIndex)
            IndustryCommentsEditor.changeRow(IndustryCommentsCurrentRow)

            IndustryCommentsEditor.changeCellValue(ISMnotelist[noteIndex].text)

            IndustryCommentsCurrentRow = IndustryCommentsInitRow
            IndustryCommentsEditor.changeRow(IndustryCommentsCurrentRow)
            break

    print("ERROR: NO MATCHING INDUSTRY")
    
print("Program executed successfully! :) Saving...")
IndustryCommentsEditor.saveToDistribution()
print("Saved!")












# PMIExcelFile.save(filename = "spreadIndustryComments.xlsx")
# print("closeColumn: {}".format(chr(closeColumn)))
# def __init__(self, name = volCalcRange[currentCloseCell].value, 
#     priority = (volCalcRange['{}{}'.format(chr(closeColumn-2),closeRow)].value), 
#     start = volCalcRange['{}{}'.format(chr(closeColumn-1),closeRow)].value, 
#     deadline = volCalcRange['{}{}'.format(chr(closeColumn+1),closeRow)].value,
#     repeating = volCalcRange[currentCloseCell].font.b, 
#     calendar = volCalcRange[currentCloseCell].font.i, 
#     color = volCalcRange[currentCloseCell].fill.start_color.index, 
#     topRow = closeRow, 
#     leftColumn = closeColumn, 
#     end = False):