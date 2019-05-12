from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Color
import datetime

def cellUpdate():
    return '{}{}'.format(chr(currentColumn),currentRow)

def txt2Cell(cellString,currentColumn,currentRow):
    currentCell = '{}{}'.format(chr(currentColumn),currentRow)
    

currentDate = datetime.date(2019, 1, 6).isoformat()

black = (0,0,0)
white = (255,255,255)
red = (255,0,0)
green = (0,255,0)
blue = (0,0,255)
grey = (128,128,128)

xTextList = 0
currentText = 0

main = load_workbook(filename = 'spreadsheet1.xlsx')
mainRange = main['Sheet1']
currentRow = 1
currentColumn = 65
# currentCell = '{}{}'.format(chr(currentColumn),currentRow)
currentCell = cellUpdate()
print("currentCell: {}".format(currentCell))

#DEFINE CELL COLORS
DARKRED = 'FFC00000'
RED = 'FFFF0000'
ORANGE = 'FFFFC000'
YELLOW = 'FFFFFF00'
GREEN = 'FF00B050'
BLUE = 'FF0070C0'
LIGHTBLUE = 'FF00B0F0'
PURPLE = 'FF7030A0'
WHITE = '00000000'
OTHERWHITE = 'FFFFFFFF'

mainRange[currentCell].value = "Hello"
print("Section {}: {}".format(currentCell,mainRange[currentCell].value))



main.save(filename = "spreadsheet1.xlsx")
# print("currentColumn: {}".format(chr(currentColumn)))

# def __init__(self, name = mainRange[currentCell].value, 
#     priority = (mainRange['{}{}'.format(chr(currentColumn-2),currentRow)].value), 
#     start = mainRange['{}{}'.format(chr(currentColumn-1),currentRow)].value, 
#     deadline = mainRange['{}{}'.format(chr(currentColumn+1),currentRow)].value,
#     repeating = mainRange[currentCell].font.b, 
#     calendar = mainRange[currentCell].font.i, 
#     color = mainRange[currentCell].fill.start_color.index, 
#     topRow = currentRow, 
#     leftColumn = currentColumn, 
#     end = False):